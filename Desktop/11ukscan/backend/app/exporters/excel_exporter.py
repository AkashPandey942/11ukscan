"""
Excel Exporter — generates a two-sheet .xlsx file from parsed statements.

Sheet 1 — "Summary":
    Professional statement summary with account info, period, totals.
    Styled with Monzo brand colours and readable typography.

Sheet 2 — "Transactions":
    Full transaction data table with header row, auto-column widths,
    alternating row colours, and frozen header.

Uses openpyxl directly for precise cell-level formatting control
that pandas alone cannot achieve.
"""

from __future__ import annotations

import logging
from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from app.core.exceptions import ExcelExportError
from app.exporters.base_exporter import AbstractExporter
from app.models.statement import ParsedStatement
from app.utils.date_utils import format_date_display, format_period

logger = logging.getLogger(__name__)

# ------------------------------------------------------------------ #
# Colour palette — Monzo brand + professional neutrals
# ------------------------------------------------------------------ #
_CORAL = "FF6B6B"          # Monzo coral/salmon
_DARK_HEADER = "1A1A2E"   # Very dark navy for primary headers
_MID_NAVY = "16213E"      # Secondary dark row
_LIGHT_BLUE = "E8F4FD"    # Alternating row light tint
_WHITE = "FFFFFF"
_CREDIT_GREEN = "00A86B"  # Green for credit amounts
_DEBIT_RED = "E53E3E"     # Red for debit amounts
_BORDER_GREY = "CCCCCC"


def _make_border(colour: str = _BORDER_GREY) -> Border:
    side = Side(border_style="thin", color=colour)
    return Border(left=side, right=side, top=side, bottom=side)


def _header_font(size: int = 11, bold: bool = True, colour: str = _WHITE) -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=colour)


def _body_font(size: int = 10, bold: bool = False, colour: str = "000000") -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=colour)


class ExcelExporter(AbstractExporter):
    """
    Exports a ParsedStatement to a professionally styled .xlsx file.

    Sheet 1 — Statement Summary: Key value pairs in a styled layout.
    Sheet 2 — Transactions:     Full table with alternating row colours,
                                frozen first row, auto-width columns,
                                and colour-coded credit/debit amounts.
    """

    extension: str = ".xlsx"

    def export(self, statement: ParsedStatement, output_path: Path) -> Path:
        """
        Write the two-sheet Excel workbook to ``output_path``.

        Args:
            statement:   Parsed statement data.
            output_path: Target .xlsx file path.

        Returns:
            The path to the written Excel file.

        Raises:
            ExcelExportError: If the file cannot be written.
        """
        logger.info(
            "Exporting %d transactions to Excel: %s",
            statement.transaction_count,
            output_path,
        )

        try:
            wb = openpyxl.Workbook()
            self._build_summary_sheet(wb, statement)
            self._build_transactions_sheet(wb, statement)
            wb.save(str(output_path))

        except ExcelExportError:
            raise
        except Exception as exc:
            logger.exception("Excel export failed: %s", exc)
            raise ExcelExportError(f"Failed to write Excel file: {exc}") from exc

        logger.info(
            "Excel export complete: %s (%d KB)",
            output_path.name,
            output_path.stat().st_size // 1024,
        )
        return output_path

    # ------------------------------------------------------------------ #
    # Sheet 1: Statement Summary
    # ------------------------------------------------------------------ #

    def _build_summary_sheet(self, wb: openpyxl.Workbook, statement: ParsedStatement) -> None:
        """Build the Statement Summary sheet (Sheet 1)."""
        ws = wb.active
        ws.title = "Summary"

        # ---- Title row ----
        ws.merge_cells("A1:B1")
        title_cell = ws["A1"]
        title_cell.value = f"BankScan — {statement.info.bank_name} Statement"
        title_cell.font = Font(name="Calibri", size=16, bold=True, color=_WHITE)
        title_cell.fill = PatternFill(fill_type="solid", fgColor=_DARK_HEADER)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36

        # ---- Section header: Account Details ----
        self._write_section_header(ws, 2, "Account Details")

        # ---- Data rows ----
        info = statement.info
        data_rows = [
            ("Account Holder", info.account_holder),
            ("Bank", info.bank_name),
            ("Account Number", info.account_number),
            ("Sort Code", info.sort_code),
            ("Statement Type", info.statement_type),
            (
                "Statement Period",
                format_period(info.period_start, info.period_end),
            ),
        ]
        row_num = 3
        for label, value in data_rows:
            self._write_data_row(ws, row_num, label, value, alternate=(row_num % 2 == 0))
            row_num += 1

        # ---- Section header: Financial Summary ----
        self._write_section_header(ws, row_num, "Financial Summary")
        row_num += 1

        closing = statement.closing_balance
        financial_rows = [
            ("Total Transactions", str(statement.transaction_count)),
            ("Total Credits", f"£{statement.total_credits:,.2f}"),
            ("Total Debits", f"£{abs(statement.total_debits):,.2f}"),
            ("Closing Balance", f"£{closing:,.2f}" if closing is not None else "N/A"),
            ("Pages Processed", str(statement.page_count)),
            (
                "Extracted On",
                format_date_display(statement.processed_at.date()),
            ),
        ]
        for label, value in financial_rows:
            self._write_data_row(ws, row_num, label, value, alternate=(row_num % 2 == 0))
            row_num += 1

        # ---- Column widths ----
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 42

    def _write_section_header(self, ws, row: int, title: str) -> None:
        """Write a visually distinct section separator row."""
        ws.merge_cells(f"A{row}:B{row}")
        cell = ws[f"A{row}"]
        cell.value = title
        cell.font = Font(name="Calibri", size=11, bold=True, color=_WHITE)
        cell.fill = PatternFill(fill_type="solid", fgColor=_CORAL)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 22

    def _write_data_row(
        self, ws, row: int, label: str, value: str, alternate: bool = False
    ) -> None:
        """Write a label-value pair row with optional alternating background."""
        bg = _LIGHT_BLUE if alternate else _WHITE

        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = _body_font(bold=True)
        label_cell.fill = PatternFill(fill_type="solid", fgColor=bg)
        label_cell.alignment = Alignment(vertical="center", indent=1)
        label_cell.border = _make_border()

        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.font = _body_font()
        value_cell.fill = PatternFill(fill_type="solid", fgColor=bg)
        value_cell.alignment = Alignment(vertical="center", indent=1)
        value_cell.border = _make_border()

        ws.row_dimensions[row].height = 20

    # ------------------------------------------------------------------ #
    # Sheet 2: Transactions
    # ------------------------------------------------------------------ #

    def _build_transactions_sheet(self, wb: openpyxl.Workbook, statement: ParsedStatement) -> None:
        """Build the full Transactions data sheet (Sheet 2)."""
        ws = wb.create_sheet(title="Transactions")

        headers = ["Date", "Description", "Amount (£)", "Balance (£)"]

        # ---- Header row ----
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = _header_font(size=11)
            cell.fill = PatternFill(fill_type="solid", fgColor=_DARK_HEADER)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _make_border()
        ws.row_dimensions[1].height = 24

        # ---- Freeze top row ----
        ws.freeze_panes = "A2"

        # ---- Transaction rows ----
        for row_idx, tx in enumerate(statement.transactions, start=2):
            alternate = row_idx % 2 == 0
            bg = _LIGHT_BLUE if alternate else _WHITE

            # Date
            date_cell = ws.cell(row=row_idx, column=1, value=tx.date.isoformat())
            date_cell.font = _body_font()
            date_cell.fill = PatternFill(fill_type="solid", fgColor=bg)
            date_cell.alignment = Alignment(horizontal="center", vertical="center")
            date_cell.border = _make_border()

            # Description
            desc_cell = ws.cell(row=row_idx, column=2, value=tx.description)
            desc_cell.font = _body_font()
            desc_cell.fill = PatternFill(fill_type="solid", fgColor=bg)
            desc_cell.alignment = Alignment(vertical="center", indent=1, wrap_text=True)
            desc_cell.border = _make_border()

            # Amount — colour coded
            amount_val = float(tx.amount)
            amount_cell = ws.cell(row=row_idx, column=3, value=amount_val)
            amount_cell.font = _body_font(
                colour=_CREDIT_GREEN if tx.is_credit else _DEBIT_RED,
                bold=True,
            )
            amount_cell.number_format = '#,##0.00'
            amount_cell.fill = PatternFill(fill_type="solid", fgColor=bg)
            amount_cell.alignment = Alignment(horizontal="right", vertical="center")
            amount_cell.border = _make_border()

            # Balance
            balance_cell = ws.cell(row=row_idx, column=4, value=float(tx.balance))
            balance_cell.font = _body_font()
            balance_cell.number_format = '#,##0.00'
            balance_cell.fill = PatternFill(fill_type="solid", fgColor=bg)
            balance_cell.alignment = Alignment(horizontal="right", vertical="center")
            balance_cell.border = _make_border()

            ws.row_dimensions[row_idx].height = 18

        # ---- Auto-size columns ----
        col_widths = [14, 55, 16, 16]
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # ---- Totals row ----
        total_row = len(statement.transactions) + 2
        ws.cell(row=total_row, column=1, value="").border = _make_border()
        ws.cell(row=total_row, column=2, value="TOTAL").font = _header_font(colour="000000")
        ws.cell(row=total_row, column=2).fill = PatternFill(fill_type="solid", fgColor="F0F0F0")
        ws.cell(row=total_row, column=2).border = _make_border()

        total_amount = ws.cell(row=total_row, column=3,
                               value=float(statement.total_credits + statement.total_debits))
        total_amount.font = _body_font(bold=True)
        total_amount.number_format = '#,##0.00'
        total_amount.fill = PatternFill(fill_type="solid", fgColor="F0F0F0")
        total_amount.border = _make_border()

        if statement.closing_balance is not None:
            closing_cell = ws.cell(row=total_row, column=4,
                                   value=float(statement.closing_balance))
            closing_cell.font = _body_font(bold=True)
            closing_cell.number_format = '#,##0.00'
            closing_cell.fill = PatternFill(fill_type="solid", fgColor="F0F0F0")
            closing_cell.border = _make_border()
