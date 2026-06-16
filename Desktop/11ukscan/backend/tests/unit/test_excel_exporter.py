"""
Unit tests for ExcelExporter.
"""

from __future__ import annotations

import pytest
import openpyxl

from app.exporters.excel_exporter import ExcelExporter


class TestExcelExporter:
    """Tests for ExcelExporter output format and sheet structure."""

    @pytest.fixture
    def exporter(self):
        return ExcelExporter()

    def test_export_creates_file(self, exporter, sample_parsed_statement, temp_output_dir):
        output_path = temp_output_dir / "test.xlsx"
        result = exporter.export(sample_parsed_statement, output_path)
        assert result.exists()

    def test_extension_is_xlsx(self):
        assert ExcelExporter.extension == ".xlsx"

    def test_workbook_has_two_sheets(self, exporter, sample_parsed_statement, temp_output_dir):
        output_path = temp_output_dir / "test.xlsx"
        exporter.export(sample_parsed_statement, output_path)

        wb = openpyxl.load_workbook(str(output_path))
        assert len(wb.sheetnames) == 2

    def test_sheet_names(self, exporter, sample_parsed_statement, temp_output_dir):
        output_path = temp_output_dir / "test.xlsx"
        exporter.export(sample_parsed_statement, output_path)

        wb = openpyxl.load_workbook(str(output_path))
        assert wb.sheetnames[0] == "Summary"
        assert wb.sheetnames[1] == "Transactions"

    def test_transactions_sheet_row_count(self, exporter, sample_parsed_statement, temp_output_dir):
        output_path = temp_output_dir / "test.xlsx"
        exporter.export(sample_parsed_statement, output_path)

        wb = openpyxl.load_workbook(str(output_path))
        ws = wb["Transactions"]
        # Header row (1) + transaction rows + totals row
        expected_rows = 1 + len(sample_parsed_statement.transactions) + 1
        assert ws.max_row == expected_rows

    def test_transactions_sheet_headers(self, exporter, sample_parsed_statement, temp_output_dir):
        output_path = temp_output_dir / "test.xlsx"
        exporter.export(sample_parsed_statement, output_path)

        wb = openpyxl.load_workbook(str(output_path))
        ws = wb["Transactions"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
        assert headers == ["Date", "Description", "Amount (£)", "Balance (£)"]

    def test_first_row_has_frozen_panes(self, exporter, sample_parsed_statement, temp_output_dir):
        output_path = temp_output_dir / "test.xlsx"
        exporter.export(sample_parsed_statement, output_path)

        wb = openpyxl.load_workbook(str(output_path))
        ws = wb["Transactions"]
        assert ws.freeze_panes == "A2"

    def test_summary_sheet_has_account_holder(self, exporter, sample_parsed_statement, temp_output_dir):
        output_path = temp_output_dir / "test.xlsx"
        exporter.export(sample_parsed_statement, output_path)

        wb = openpyxl.load_workbook(str(output_path))
        ws = wb["Summary"]
        # Collect all cell values
        all_values = [ws.cell(row=r, column=c).value
                      for r in range(1, ws.max_row + 1)
                      for c in range(1, 3)]
        assert sample_parsed_statement.info.account_holder in all_values
